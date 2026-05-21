"""
RFP Generator Backend Server
Runs locally. Dashboard calls this to:
1. Upload file → AI generates RFP data
2. Download → Produces formatted Excel using template
"""
import json, os, sys, shutil, httpx, anthropic
from pathlib import Path
from copy import copy
from datetime import datetime
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

app = Flask(__name__)
CORS(app)

# Load .env
env_file = Path(__file__).parent / '.env'
if env_file.exists():
    for line in env_file.read_text().strip().split('\n'):
        if '=' in line and not line.startswith('#'):
            k, v = line.split('=', 1)
            os.environ.setdefault(k.strip(), v.strip())

API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
MODEL = "bedrock.anthropic.claude-sonnet-4-6"
TEMPLATE = r"C:\Users\igarg015\Downloads\Shared_Viatris_ Temporary Labor_Agency_RFX (1).xlsx"
OUTPUT_DIR = Path(__file__).parent / "output"
OUTPUT_DIR.mkdir(exist_ok=True)

SYSTEM = """You are a Senior Procurement RFP Specialist writing TEMPORARY LABOR / CONTINGENT WORKFORCE RFPs for staffing agencies.

The output is ALWAYS a Temporary Labor RFP sent to STAFFING AGENCIES. The input tells you about the CLIENT. You create an RFP asking staffing suppliers to provide temp workers to that client.

Output valid JSON only (no markdown fences):
{
  "projectSetup": {"clientName":"","industry":"","description":"2-sentence company description","contactName":"","contactEmail":"","timeline":{"issueDate":"","questionDeadline":"","proposalDue":""},"duration":"","currency":"USD","countries":"","units":"","headcount":"","spend":"","paymentTerms":"Net 60"},
  "requirements": {"overview":"program overview paragraph","items":[{"category":"All","subcategory":"","text":""}]},
  "questionnaire": {"sections":[{"name":"","questions":[{"text":"","type":"Text"}]}]},
  "slas": [{"name":"","target":"","description":"","frequency":"Monthly","source":""}],
  "bidSheets": {"instructions":"","roles":[{"id":"","region":"APAC|EMEA|LATAM|NA","family":"","title":"","desc":"","country":"","site":"","demand":"","rate":""}]},
  "markup": {"rows":[{"region":"","location":"","country":"","family":"","demand":""}]},
  "discounts": {"conversion":[],"rebate":[],"tenure":[]},
  "conversion": {"fxRates":[{"location":"","country":"","currency":"","rate":""}]}
}

REQUIREMENTS must cover: Workforce Readiness, Staffing Controls, Compliance, Service Delivery, Reporting, Health/Safety, Commercial Management, Transition.
QUESTIONNAIRE sections: Account Management, Compliance & Risk, Fulfillment, Geographic Coverage, Pricing, Recruitment, Service Delivery, Technology & Reporting.
SLAs: Fill Rate >=90%, Time-to-Fill <=14 days, No-Show <=3%, Invoice Accuracy >=98%, Replacement 24-48hrs, Onboarding 3 days, Turnover <=7%, Quality >65%, Retention >=80%.

Generate 30 requirements, 30 questions, 15 SLAs, roles from client needs with market pay rates. Keep text concise (1-2 sentences).

CRITICAL REGION MAPPING FOR BID SHEETS:
- APAC = India, China, Japan, Thailand, Singapore, Australia, Philippines, Vietnam, Indonesia, Malaysia, Korea, Bangladesh, Pakistan
- EMEA = UK, Germany, France, Ireland, Switzerland, Egypt, South Africa, Italy, Spain, Netherlands, UAE, Saudi Arabia, Nigeria
- LATAM = Brazil, Mexico, Argentina, Colombia, Chile, Peru
- NA = United States, Canada, Puerto Rico

Map each role to the CORRECT region based on its country. Do NOT put India roles in NA.

Output ONLY valid JSON."""


def read_file(filepath):
    ext = Path(filepath).suffix.lower()
    if ext in ('.xlsx', '.xls'):
        wb = load_workbook(filepath, data_only=True)
        parts = []
        for name in wb.sheetnames:
            ws = wb[name]
            parts.append(f'\n=== {name} ===')
            for row in ws.iter_rows(values_only=True):
                vals = [str(c).strip() for c in row if c]
                if vals: parts.append(' | '.join(vals))
        return '\n'.join(parts)
    elif ext == '.docx':
        try: from docx import Document
        except: os.system("pip install python-docx -q"); from docx import Document
        return '\n'.join(p.text for p in Document(filepath).paragraphs if p.text.strip())
    elif ext == '.pdf':
        try: import PyPDF2
        except: os.system("pip install PyPDF2 -q"); import PyPDF2
        with open(filepath,'rb') as f:
            return '\n'.join(p.extract_text() or '' for p in PyPDF2.PdfReader(f).pages)
    else:
        with open(filepath, 'r', encoding='utf-8', errors='ignore') as f: return f.read()


def call_ai(content):
    if len(content) > 60000: content = content[:60000]
    client = anthropic.Anthropic(api_key=API_KEY, http_client=httpx.Client(verify=False))
    r = client.messages.create(model=MODEL, max_tokens=8000, system=SYSTEM,
        messages=[{"role":"user","content":"Generate Temp Labor RFP JSON. Be concise, 1 sentence per item. Complete valid JSON under 7000 tokens:\n\n"+content}])
    raw = r.content[0].text.strip()
    if raw.startswith('```'): raw = raw.split('\n',1)[1]
    if raw.endswith('```'): raw = raw[:-3]
    if raw.startswith('json'): raw = raw[4:]
    raw = raw.strip()
    # Fix truncated JSON
    opens = raw.count('{') - raw.count('}')
    opens2 = raw.count('[') - raw.count(']')
    if opens > 0 or opens2 > 0:
        last_good = max(raw.rfind('}'), raw.rfind(']'))
        if last_good > len(raw)//2:
            raw = raw[:last_good+1]
            opens = raw.count('{') - raw.count('}')
            opens2 = raw.count('[') - raw.count(']')
            raw += ']'*opens2 + '}'*opens
    return json.loads(raw)


def build_excel(D):
    """Build Excel using template formatting, filled with client data."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # for custom tabs

    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    ps = D.get('projectSetup', {})
    client = ps.get('clientName', 'Client')
    safe_client = client.replace(' ','_').replace('/','_')
    output_path = str(OUTPUT_DIR / f"{safe_client}_Temp_Labor_RFP_{ts}.xlsx")

    shutil.copy2(TEMPLATE, output_path)
    wb = load_workbook(output_path)

    # Remove all logos
    for sn in wb.sheetnames:
        wb[sn]._images = []

    # Unmerge all merged cells so we can write freely
    for sn in wb.sheetnames:
        ws = wb[sn]
        for mr in list(ws.merged_cells.ranges):
            ws.unmerge_cells(str(mr))

    # Replace Viatris with client name everywhere (keeps formatting!)
    for sn in wb.sheetnames:
        ws = wb[sn]
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                if isinstance(cell, MergedCell): continue
                if cell.value and isinstance(cell.value, str) and 'Viatris' in cell.value:
                    cell.value = cell.value.replace('Viatris', client)

    # Clear only DATA rows (preserves headers, instructions, formatting)
    # Tab 3: data starts row 8
    ws = wb['3. Requirements']
    for r in range(8, ws.max_row+1):
        for c in range(2, 8):
            cell = ws.cell(row=r, column=c)
            if not isinstance(cell, MergedCell): cell.value = None
    # Tab 4: data starts row 8
    ws = wb['4. Questionnaire']
    for r in range(8, ws.max_row+1):
        for c in range(2, 6):
            cell = ws.cell(row=r, column=c)
            if not isinstance(cell, MergedCell): cell.value = None
    # Tab 5: data starts row 8
    ws = wb['5. Service Level Agreements']
    for r in range(8, ws.max_row+1):
        for c in range(2, 10):
            cell = ws.cell(row=r, column=c)
            if not isinstance(cell, MergedCell): cell.value = None
    # Tabs 6a-d: data starts after header row (find RFP ID row + 2)
    for sn in ['6.a Bid Sheet_APAC','6.b Bid Sheet_EMEA','6.c Bid Sheet_LATAM','6.d Bid Sheet_NA']:
        ws = wb[sn]
        start = 15  # default
        for r in range(1, 20):
            if ws.cell(row=r, column=2).value and 'RFP ID' in str(ws.cell(row=r, column=2).value):
                start = r + 2; break
        for r in range(start, ws.max_row+1):
            for c in range(2, 14):
                cell = ws.cell(row=r, column=c)
                if not isinstance(cell, MergedCell): cell.value = None
    # Tab 7: data starts after header
    ws = wb['7. Bid Sheet_Mark-Up']
    for r in range(14, ws.max_row+1):
        for c in range(1, 10):
            cell = ws.cell(row=r, column=c)
            if not isinstance(cell, MergedCell): cell.value = None
    # Tab 9: data from row 6
    ws = wb['9. Conversion Rates']
    for r in range(6, ws.max_row+1):
        for c in range(3, 7):
            cell = ws.cell(row=r, column=c)
            if not isinstance(cell, MergedCell): cell.value = None

    # Template formatting preserved — just write data values into cells
    tl = ps.get('timeline', {})

    # TAB 3: Write requirements data into existing formatted rows
    ws = wb['3. Requirements']
    for i, req in enumerate(D.get('requirements',{}).get('items',[])):
        r = 8+i
        ws.cell(row=r,column=2).value = i+1
        ws.cell(row=r,column=3).value = req.get('category','All')
        ws.cell(row=r,column=4).value = req.get('subcategory','')
        ws.cell(row=r,column=5).value = req.get('text','')

    # TAB 4: Write questionnaire
    ws = wb['4. Questionnaire']
    qi=0
    for sec in D.get('questionnaire',{}).get('sections',[]):
        for q in sec.get('questions',[]):
            qi+=1; r=7+qi
            ws.cell(row=r,column=2).value = qi
            ws.cell(row=r,column=3).value = sec['name']
            ws.cell(row=r,column=4).value = q.get('text','')

    # TAB 5: Write SLAs
    ws = wb['5. Service Level Agreements']
    for i, s in enumerate(D.get('slas',[])):
        r=8+i
        ws.cell(row=r,column=2).value = s.get('name','')
        ws.cell(row=r,column=3).value = s.get('description','')
        ws.cell(row=r,column=5).value = s.get('target','')
        ws.cell(row=r,column=6).value = s.get('frequency','Monthly')
        ws.cell(row=r,column=7).value = s.get('source','')

    # TABs 6a-d: Write bid sheet roles
    rmap = {'APAC':'6.a Bid Sheet_APAC','EMEA':'6.b Bid Sheet_EMEA','LATAM':'6.c Bid Sheet_LATAM','NA':'6.d Bid Sheet_NA'}
    roles = D.get('bidSheets',{}).get('roles',[])
    for region, sn in rmap.items():
        ws = wb[sn]
        # Find the data start row (row after "Example>" or after header)
        start = 15
        for r in range(1, 20):
            v = ws.cell(row=r, column=2).value
            if v and 'RFP ID' in str(v): start = r + 2; break
        rr=[x for x in roles if x.get('region')==region]
        for i,role in enumerate(rr):
            r=start+i
            ws.cell(row=r,column=2).value=role.get('id','')
            ws.cell(row=r,column=3).value=role.get('family','')
            ws.cell(row=r,column=4).value=role.get('title','')
            ws.cell(row=r,column=5).value=role.get('desc','')
            ws.cell(row=r,column=6).value=role.get('country','')
            ws.cell(row=r,column=7).value=role.get('site','')
            ws.cell(row=r,column=8).value=role.get('demand','')
            ws.cell(row=r,column=9).value=role.get('rate','')

    # TAB 7: Write mark-up rows
    ws = wb['7. Bid Sheet_Mark-Up']
    start = 14
    for r in range(1, 20):
        v = ws.cell(row=r, column=2).value
        if v and 'Region' in str(v): start = r + 2; break
    for i,m in enumerate(D.get('markup',{}).get('rows',[])):
        r=start+i
        ws.cell(row=r,column=2).value=m.get('region','')
        ws.cell(row=r,column=3).value=m.get('location','')
        ws.cell(row=r,column=4).value=m.get('country','')
        ws.cell(row=r,column=5).value=m.get('family','')
        ws.cell(row=r,column=6).value=m.get('demand','')

    # TAB 9: Write conversion rates
    ws = wb['9. Conversion Rates']
    for i,fx in enumerate(D.get('conversion',{}).get('fxRates',[])):
        r=6+i
        ws.cell(row=r,column=3).value=fx.get('location','')
        ws.cell(row=r,column=4).value=fx.get('country','')
        ws.cell(row=r,column=5).value=fx.get('currency','')
        try: ws.cell(row=r,column=6).value=float(fx.get('rate',0))
        except: ws.cell(row=r,column=6).value=fx.get('rate','')

    # Handle any custom tabs from dashboard
    custom_tabs = D.get('customTabs', [])
    for ct in custom_tabs:
        sn = ct['name'][:31]
        if sn not in wb.sheetnames:
            wb.create_sheet(sn)
        ws = wb[sn]
        cols = ct.get('cols', [])
        # Title
        style_title(ws, 1, 1, f"{client} — {ct['name']}")
        # Headers with styling
        if cols:
            for ci, col in enumerate(cols):
                c = ws.cell(row=3, column=ci+1); c.value=col; c.font=header_font; c.fill=header_fill; c.alignment=Alignment(wrap_text=True)
        # Data rows
        for ri, row in enumerate(ct.get('rows', [])):
            for ci, val in enumerate(row):
                style_data_cell(ws, 4+ri, ci+1, val)
        # Set column widths
        for ci in range(len(cols)):
            ws.column_dimensions[chr(65+ci)].width = 20

    wb.save(output_path)
    return output_path


# ===== API ENDPOINTS =====

@app.route('/api/generate', methods=['POST'])
def generate():
    """Upload file → AI generates RFP data → returns JSON to dashboard."""
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    file = request.files['file']
    # Save temp
    temp_path = str(OUTPUT_DIR / f"_temp_{file.filename}")
    file.save(temp_path)
    try:
        content = read_file(temp_path)
        data = call_ai(content)
        return jsonify(data)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        try: os.remove(temp_path)
        except: pass


@app.route('/api/export', methods=['POST'])
def export():
    """Takes dashboard data (with edits) → produces formatted Excel → returns file."""
    data = request.get_json()
    if not data:
        return jsonify({'error': 'No data provided'}), 400
    try:
        path = build_excel(data)
        return send_file(path, as_attachment=True, download_name=os.path.basename(path))
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/health', methods=['GET'])
def health():
    return jsonify({'status': 'ok', 'template_exists': os.path.exists(TEMPLATE)})


if __name__ == '__main__':
    print("="*50)
    print("  RFP Generator Server")
    print("  http://localhost:5000")
    print("="*50)
    app.run(port=5000, debug=False)
