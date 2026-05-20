"""
RFP Generator - One command, one output.
Upload any client file → AI generates complete RFP → Outputs Excel directly.

Usage: python run_rfp.py <client_file>
"""
import json, sys, os, shutil, httpx, anthropic
from pathlib import Path

# Load .env file if it exists
env_file = Path(__file__).parent / '.env'
if env_file.exists():
    for line in env_file.read_text().strip().split('\n'):
        if '=' in line and not line.startswith('#'):
            k, v = line.split('=', 1)
            os.environ.setdefault(k.strip(), v.strip())
from copy import copy
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
MODEL = "bedrock.anthropic.claude-sonnet-4-6"
TEMPLATE = r"C:\Users\igarg015\Downloads\Shared_Viatris_ Temporary Labor_Agency_RFX (1).xlsx"
OUTPUT_DIR = r"c:\Isha\Claude\Temp labor\output"

SYSTEM = """You are a Senior Procurement RFP Specialist writing TEMPORARY LABOR / CONTINGENT WORKFORCE RFPs for staffing agencies.

CRITICAL RULE: The output is ALWAYS a Temporary Labor RFP sent to STAFFING AGENCIES. It is NOT a copy of the input document. The input document just tells you about the CLIENT (their industry, locations, needs). You then create a Temp Labor RFP that asks staffing suppliers to provide temporary workers to that client.

WHAT A TEMP LABOR RFP CONTAINS:
- Requirements for STAFFING AGENCIES (background checks, onboarding, timekeeping, compliance, co-employment safeguards, reporting)
- Questionnaire asking STAFFING AGENCIES about their capabilities (account management, fulfillment, technology/VMS, geographic coverage)
- SLAs for STAFFING PERFORMANCE (fill rate, time-to-fill, no-show rate, replacement turnaround, invoice accuracy, retention)
- Bid sheets listing TEMP WORKER ROLES the client needs filled (with pay rates for the staffing agency to price)
- Mark-up structure for AGENCY FEES on top of worker pay
- Discounts/rebates the AGENCY offers for volume

From the input document, EXTRACT:
- Client company name and industry
- Locations/countries where they operate
- Types of roles/skills they need (translate these into temp labor job categories)
- Any timeline, headcount, or budget info
- Compliance/regulatory context (pharma=GMP, finance=SOX, etc.)

Then GENERATE a complete Temp Labor RFP with content appropriate for that client's industry.

Output valid JSON only (no markdown fences):
{
  "projectSetup": {"clientName":"","industry":"","rfpTitle":"Temporary Labor Global Agency RFP","description":"company description","contactName":"","contactEmail":"","timeline":{"issueDate":"","questionDeadline":"","proposalDue":""},"duration":"3 years","currency":"USD","countries":"comma-separated","units":"business units","headcount":"","spend":"","paymentTerms":"Net 60"},
  "requirements": {"overview":"program overview paragraph","items":[{"category":"All|Operations|Sterile|Commercial","subcategory":"Workforce Readiness|Staffing Controls|Compliance|Reporting|Safety|Commercial","text":"requirement for the staffing agency"}]},
  "questionnaire": {"sections":[{"name":"section name","questions":[{"text":"question for staffing agency","type":"Text"}]}]},
  "slas": [{"name":"KPI name","target":"e.g. >=90%","description":"what it measures","frequency":"Monthly","source":"VMS/reporting"}],
  "bidSheets": {"instructions":"instructions for staffing agencies on how to fill rates","roles":[{"id":"XX0001","region":"APAC|EMEA|LATAM|NA","family":"job family","title":"temp worker role title","desc":"role description","country":"country","site":"city","demand":"headcount","rate":"USD/hr pay rate"}]},
  "markup": {"rows":[{"region":"","location":"city, country","country":"","family":"job family","demand":"number"}]},
  "discounts": {"conversion":[{"low":"0","high":"30","pct":"15"}],"rebate":[],"tenure":[]},
  "conversion": {"fxRates":[{"location":"city","country":"country","currency":"Currency (CODE)","rate":"rate to 1 USD"}]}
}

REQUIREMENTS must cover these temp labor categories:
- Workforce Readiness & Onboarding (screening, training, documentation)
- Workforce Management & Staffing Controls (timesheets, retention, replacements)
- Service Model & Compliance (co-employment, worker classification, badges)
- Service Delivery (fill rates, no-shows, attrition)
- Reporting & Performance (monthly reports, KPIs, data privacy)
- Health, Safety & Security
- Commercial Management (invoicing, payment terms, OT billing)
- Change & Transition (offboarding, supplier transition)

QUESTIONNAIRE sections must be:
- Account Management, Compliance & Risk, Fulfillment Capability, Geographic Coverage, Pricing & Commercials, Recruitment Process, Service Delivery, Technology & Reporting, Diversity & Inclusion

SLAs must include: Fill Rate (>=90%), Time-to-Fill (<=14 days), No-Show Rate (<=3%), Invoice Accuracy (>=98%), Replacement Turnaround (24-48hrs), Onboarding Time (3 days), Early Turnover (<=7%), Candidate Quality (>65%), Retention (>=80%), Customer Support Response (1 day)

Generate 30 requirements, 30 questions, 15 SLAs, roles based on client needs, mark-up rows per location/family. Keep all text concise (1-2 sentences). Output ONLY valid JSON."""


def read_file(path):
    ext = os.path.splitext(path)[1].lower()
    if ext in ('.xlsx', '.xls'):
        from openpyxl import load_workbook as lw
        wb = lw(path, data_only=True)
        parts = []
        for name in wb.sheetnames:
            ws = wb[name]
            parts.append(f'\n=== {name} ===')
            for row in ws.iter_rows(values_only=True):
                vals = [str(c).strip() for c in row if c]
                if vals: parts.append(' | '.join(vals))
        return '\n'.join(parts)
    elif ext == '.docx':
        try: from docx import Document
        except: os.system("pip install python-docx -q"); from docx import Document
        return '\n'.join(p.text for p in Document(path).paragraphs if p.text.strip())
    elif ext == '.pdf':
        try: import PyPDF2
        except: os.system("pip install PyPDF2 -q"); import PyPDF2
        with open(path,'rb') as f:
            return '\n'.join(p.extract_text() or '' for p in PyPDF2.PdfReader(f).pages)
    else:
        with open(path, 'r', encoding='utf-8', errors='ignore') as f: return f.read()


def call_ai(content):
    if len(content) > 60000: content = content[:60000]
    client = anthropic.Anthropic(api_key=API_KEY, http_client=httpx.Client(verify=False))
    print("  Calling AI...")
    r = client.messages.create(model=MODEL, max_tokens=8000, system=SYSTEM,
        messages=[{"role":"user","content":"Generate a Temp Labor RFP JSON for the client described below. IMPORTANT: In projectSetup.description, write a 2-sentence description of THIS specific client (their industry, what they do). In projectSetup.countries, list the actual countries where they operate. In requirements.overview, write a paragraph about THIS client's temp labor program goals. Be concise: 30 requirements, 30 questions, 15 SLAs, keep text to 1 sentence each. Complete valid JSON under 7000 tokens:\n\n"+content}])
    raw = r.content[0].text.strip()
    if raw.startswith('```'): raw = raw.split('\n',1)[1]
    if raw.endswith('```'): raw = raw[:-3]
    if raw.startswith('json'): raw = raw[4:]
    raw = raw.strip()
    # Fix truncated JSON by closing open brackets
    opens = raw.count('{') - raw.count('}')
    opens2 = raw.count('[') - raw.count(']')
    if opens > 0 or opens2 > 0:
        # Try to find last complete object and truncate there
        last_good = max(raw.rfind('}'), raw.rfind(']'))
        if last_good > len(raw)//2:
            raw = raw[:last_good+1]
            # Close remaining brackets
            opens = raw.count('{') - raw.count('}')
            opens2 = raw.count('[') - raw.count(']')
            raw += ']'*opens2 + '}'*opens
    return json.loads(raw)


def safe_write(ws, row, col, val):
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        for mr in list(ws.merged_cells.ranges):
            if cell.coordinate in mr: ws.unmerge_cells(str(mr)); break
        cell = ws.cell(row=row, column=col)
    cell.value = val

def clear_rows(ws, start, cols):
    for r in range(start, ws.max_row+1):
        for c in range(1, cols+1):
            cell = ws.cell(row=r, column=c)
            if not isinstance(cell, MergedCell): cell.value = None

def find_row(ws, text, col=2):
    for r in range(1, min(ws.max_row+1, 30)):
        v = ws.cell(row=r, column=col).value
        if v and text in str(v): return r
    return None


def build_excel(D, output_path):
    shutil.copy2(TEMPLATE, output_path)
    wb = load_workbook(output_path)
    ps = D.get('projectSetup', {})
    client = ps.get('clientName', 'Client')
    tl = ps.get('timeline', {})

    # STEP 1: REMOVE ALL IMAGES/LOGOS (Viatris branding)
    for sn in wb.sheetnames:
        wb[sn]._images = []

    # STEP 2: UNMERGE ALL CELLS (so we can write freely)
    for sn in wb.sheetnames:
        ws = wb[sn]
        for mr in list(ws.merged_cells.ranges):
            ws.unmerge_cells(str(mr))

    # STEP 3: CLEAR ALL CELL VALUES (keep column widths, row heights, fill colors)
    for sn in wb.sheetnames:
        ws = wb[sn]
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                if isinstance(cell, MergedCell): continue
                cell.value = None

    # STEP 4: WRITE ALL CONTENT FRESH (no old data remains)

    # --- TAB 1: Cover Letter ---
    ws = wb['1. Cover Letter']
    ws.cell(row=2, column=2).value = f"{client} Temporary Labor Global Agency RFP"
    ws.cell(row=4, column=2).value = f"About {client}"
    ws.cell(row=4, column=3).value = ps.get('description', '')
    ws.cell(row=6, column=2).value = "RFP Objective"
    ws.cell(row=6, column=3).value = f"This RFP is being issued by {client} to Temporary Labor agencies. " + D.get('requirements',{}).get('overview','')
    ws.cell(row=8, column=2).value = "RFP Scope"
    ws.cell(row=8, column=3).value = f"This RFP covers Temporary Labor requirements in: {ps.get('countries', 'Global')}"
    ws.cell(row=9, column=3).value = f"Business Units: {ps.get('units', '')}"
    ws.cell(row=11, column=2).value = "Confidentiality"
    ws.cell(row=11, column=3).value = f"All information in this document is the confidential property of {client}."
    ws.cell(row=13, column=2).value = "RFP Process Timeline"
    ws.cell(row=14, column=3).value = f"- RFP Release: {tl.get('issueDate', 'TBD')}"
    ws.cell(row=15, column=3).value = f"- Vendor Questions Due: {tl.get('questionDeadline', 'TBD')}"
    ws.cell(row=16, column=3).value = f"- Submission Deadline: {tl.get('proposalDue', 'TBD')}"
    ws.cell(row=18, column=2).value = "Bid Instructions"
    ws.cell(row=18, column=3).value = D.get('bidSheets',{}).get('instructions', 'Complete all tabs per instructions.')
    ws.cell(row=20, column=2).value = f"{client} Point of Contact"
    ws.cell(row=21, column=2).value = f"Name: {ps.get('contactName', '')}"
    ws.cell(row=22, column=2).value = f"Email: {ps.get('contactEmail', '')}"
    ws.cell(row=23, column=2).value = f"Payment Terms: {ps.get('paymentTerms', 'Net 60')}"

    # --- TAB 2: Contents ---
    ws = wb['2. Contents & Response Summary']
    ws.cell(row=1, column=4).value = f"{client} Temporary Labor RFP"
    ws.cell(row=2, column=4).value = "Contents & Response Summary"
    ws.cell(row=4, column=2).value = "RFP Workbook Contents"
    ws.cell(row=5, column=3).value = "#"; ws.cell(row=5, column=4).value = "Tab"; ws.cell(row=5, column=5).value = "Description"; ws.cell(row=5, column=6).value = "Action Required"
    tabs_info = [("1","Cover Letter","RFP overview","Review"),("2","Contents","Workbook index","Track progress"),("3","Requirements","Business requirements","Yes/No/Partial + Comments"),("4","Questionnaire","Capability questions","Narrative responses"),("5","SLAs","KPIs and targets","Comply Y/N + Comments"),("6a-d","Bid Sheets","Roles and pay rates by region","Proposed rates"),("7","Mark-Up","Mark-up percentages","Standard + OT %"),("8","Discounts","Commercial incentives","Tiered pricing"),("9","Conversion Rates","Currency reference","Reference only")]
    for i, (num, tab, desc, action) in enumerate(tabs_info):
        r = 6 + i
        ws.cell(row=r, column=3).value = num; ws.cell(row=r, column=4).value = tab
        ws.cell(row=r, column=5).value = desc; ws.cell(row=r, column=6).value = action

    # --- TAB 3: Requirements ---
    ws = wb['3. Requirements']
    ws.cell(row=1, column=3).value = f"{client} Temporary Labor RFP"
    ws.cell(row=2, column=3).value = "Business Requirements"
    ws.cell(row=4, column=2).value = "Instructions:"
    ws.cell(row=5, column=2).value = "Please provide a response by selecting Yes, No, or Partial from the dropdown. Add comments where needed."
    ws.cell(row=7, column=2).value = "#"; ws.cell(row=7, column=3).value = "Category"; ws.cell(row=7, column=4).value = "Sub-Category"; ws.cell(row=7, column=5).value = "Requirement"; ws.cell(row=7, column=6).value = "Response"; ws.cell(row=7, column=7).value = "Comment"
    items = D.get('requirements',{}).get('items',[])
    for i, req in enumerate(items):
        r = 8 + i
        ws.cell(row=r, column=2).value = i+1
        ws.cell(row=r, column=3).value = req.get('category','All')
        ws.cell(row=r, column=4).value = req.get('subcategory','')
        ws.cell(row=r, column=5).value = req.get('text','')

    # --- TAB 4: Questionnaire ---
    ws = wb['4. Questionnaire']
    ws.cell(row=1, column=3).value = f"{client} Temporary Labor RFP"
    ws.cell(row=2, column=3).value = "Questionnaire"
    ws.cell(row=4, column=2).value = "Instructions:"
    ws.cell(row=5, column=2).value = "Please answer each question. All questions are required. Provide responses in column E."
    ws.cell(row=7, column=2).value = "#"; ws.cell(row=7, column=3).value = "Category"; ws.cell(row=7, column=4).value = "Question"; ws.cell(row=7, column=5).value = "Response"
    questions = []
    for sec in D.get('questionnaire',{}).get('sections',[]):
        for q in sec.get('questions',[]): questions.append({'cat':sec['name'],'text':q.get('text','')})
    for i, q in enumerate(questions):
        r = 8 + i
        ws.cell(row=r, column=2).value = i+1
        ws.cell(row=r, column=3).value = q['cat']
        ws.cell(row=r, column=4).value = q['text']

    # --- TAB 5: SLAs ---
    ws = wb['5. Service Level Agreements']
    ws.cell(row=1, column=3).value = f"{client} Temporary Labor RFP"
    ws.cell(row=2, column=3).value = "Service Level Agreements (SLA)"
    ws.cell(row=4, column=2).value = "Instructions:"
    ws.cell(row=5, column=2).value = "Review SLAs and confirm compliance. Provide rationale if unable to comply."
    ws.cell(row=7, column=2).value = "KPI"; ws.cell(row=7, column=3).value = "Requirement Definition"; ws.cell(row=7, column=4).value = "Measurement Calculation"; ws.cell(row=7, column=5).value = "Target"; ws.cell(row=7, column=6).value = "Period"; ws.cell(row=7, column=7).value = "Source"; ws.cell(row=7, column=8).value = "Willing to Comply (Yes/No)"; ws.cell(row=7, column=9).value = "Comments"
    slas = D.get('slas',[])
    for i, s in enumerate(slas):
        r = 8 + i
        ws.cell(row=r, column=2).value = s.get('name','')
        ws.cell(row=r, column=3).value = s.get('description','')
        ws.cell(row=r, column=5).value = s.get('target','')
        ws.cell(row=r, column=6).value = s.get('frequency','Monthly')
        ws.cell(row=r, column=7).value = s.get('source','')

    # --- TABs 6a-d: Bid Sheets ---
    rmap = {'APAC':'6.a Bid Sheet_APAC','EMEA':'6.b Bid Sheet_EMEA','LATAM':'6.c Bid Sheet_LATAM','NA':'6.d Bid Sheet_NA'}
    roles = D.get('bidSheets',{}).get('roles',[])
    for region, sn in rmap.items():
        ws = wb[sn]
        ws.cell(row=1, column=3).value = f"{client} Temporary Labor RFP"
        ws.cell(row=2, column=3).value = f"Bid Sheet: {region}"
        ws.cell(row=4, column=2).value = "Please use USD/hour for all rates."
        ws.cell(row=5, column=2).value = "Fill ONLY Base Pay Rates here. Mark-Up values go in Tab 7 only."
        ws.cell(row=7, column=2).value = "Rate Definitions:"
        ws.cell(row=8, column=2).value = "Pay Rate - Standard (USD): Salary paid directly to the temporary worker"
        ws.cell(row=9, column=2).value = "Pay Rate - Overtime (USD): Overtime salary paid to the temporary worker"
        # Header row
        headers = ['RFP ID','Job Family','Job Title','Job Description','Location - Country','Location - Site','Projected Demand','Pay Rate Standard (USD)','Proposed Pay Rate (USD)','OT Multiplier','Bidding? (Yes/No)','Comments']
        for ci, h in enumerate(headers):
            ws.cell(row=11, column=2+ci).value = h
        # Data
        rr = [x for x in roles if x.get('region')==region]
        for i, role in enumerate(rr):
            r = 12 + i
            ws.cell(row=r, column=2).value = role.get('id','')
            ws.cell(row=r, column=3).value = role.get('family','')
            ws.cell(row=r, column=4).value = role.get('title','')
            ws.cell(row=r, column=5).value = role.get('desc','')
            ws.cell(row=r, column=6).value = role.get('country','')
            ws.cell(row=r, column=7).value = role.get('site','')
            ws.cell(row=r, column=8).value = role.get('demand','')
            ws.cell(row=r, column=9).value = role.get('rate','')

    # --- TAB 7: Mark-Up ---
    ws = wb['7. Bid Sheet_Mark-Up']
    ws.cell(row=1, column=3).value = f"{client} Temporary Labor RFP"
    ws.cell(row=2, column=3).value = "Bid Sheet: Mark-up"
    ws.cell(row=4, column=2).value = "Provide ALL Mark-Up values in this tab only. Do NOT include in other tabs."
    ws.cell(row=6, column=2).value = "Mark up - Standard (%): Includes statutory costs, fringe benefits, operating expenses, and service fees"
    ws.cell(row=7, column=2).value = "Mark up - Overtime (%): Overtime mark-up applied to the worker pay rate"
    headers = ['Region','Location (City, Country)','Country','Job Family','Projected Demand','Mark-up Standard (%)','Mark-up Overtime (%)','Comments']
    for ci, h in enumerate(headers):
        ws.cell(row=9, column=2+ci).value = h
    mus = D.get('markup',{}).get('rows',[])
    for i, m in enumerate(mus):
        r = 10 + i
        ws.cell(row=r, column=2).value = m.get('region','')
        ws.cell(row=r, column=3).value = m.get('location','')
        ws.cell(row=r, column=4).value = m.get('country','')
        ws.cell(row=r, column=5).value = m.get('family','')
        ws.cell(row=r, column=6).value = m.get('demand','')

    # --- TAB 8: Discounts ---
    ws = wb['8. Discounts']
    ws.cell(row=1, column=3).value = f"{client} Temporary Labor RFP"
    ws.cell(row=2, column=3).value = "Discounts"
    ws.cell(row=4, column=2).value = "Instructions: Please provide responses to each question below."
    r = 6
    ws.cell(row=r, column=2).value = "1"; r+=1
    ws.cell(row=r, column=2).value = f"{client} expects a conversion fee of $0 unless stated otherwise."; r+=2
    ws.cell(row=r, column=3).value = "Range - Low End (Days)"; ws.cell(row=r, column=4).value = "Range - High End (Days)"; ws.cell(row=r, column=5).value = "Conversion Fee (%)"; r+=1
    for d in D.get('discounts',{}).get('conversion',[]):
        ws.cell(row=r, column=3).value = d.get('low',''); ws.cell(row=r, column=4).value = d.get('high',''); ws.cell(row=r, column=5).value = d.get('pct',''); r+=1
    r+=1; ws.cell(row=r, column=2).value = "2"; r+=1
    ws.cell(row=r, column=2).value = "Describe any volume rebates."; r+=2
    ws.cell(row=r, column=3).value = "Spend Low"; ws.cell(row=r, column=4).value = "Spend High"; ws.cell(row=r, column=5).value = "Rebate %"; r+=1
    for d in D.get('discounts',{}).get('rebate',[]):
        ws.cell(row=r, column=3).value = d.get('low',''); ws.cell(row=r, column=4).value = d.get('high',''); ws.cell(row=r, column=5).value = d.get('pct',''); r+=1
    r+=1; ws.cell(row=r, column=2).value = "3"; r+=1
    ws.cell(row=r, column=2).value = "Describe any tenure discounts."; r+=2
    ws.cell(row=r, column=3).value = "Low (Months)"; ws.cell(row=r, column=4).value = "High (Months)"; ws.cell(row=r, column=5).value = "Discount %"; r+=1
    for d in D.get('discounts',{}).get('tenure',[]):
        ws.cell(row=r, column=3).value = d.get('low',''); ws.cell(row=r, column=4).value = d.get('high',''); ws.cell(row=r, column=5).value = d.get('pct',''); r+=1

    # --- TAB 9: Conversion Rates ---
    ws = wb['9. Conversion Rates']
    ws.cell(row=1, column=3).value = f"{client} Temporary Labor RFP"
    ws.cell(row=2, column=3).value = "USD Conversion Rate Table"
    ws.cell(row=4, column=3).value = "Location"; ws.cell(row=4, column=4).value = "Country"; ws.cell(row=4, column=5).value = "Currency"; ws.cell(row=4, column=6).value = "Conversion rate to $1.00 USD"
    fx = D.get('conversion',{}).get('fxRates',[])
    for i, c in enumerate(fx):
        r = 5 + i
        ws.cell(row=r, column=3).value = c.get('location','')
        ws.cell(row=r, column=4).value = c.get('country','')
        ws.cell(row=r, column=5).value = c.get('currency','')
        try: ws.cell(row=r, column=6).value = float(c.get('rate',0))
        except: ws.cell(row=r, column=6).value = c.get('rate','')

    wb.save(output_path)


def main():
    if len(sys.argv) < 2:
        print("Usage: python run_rfp.py <client_file>")
        print("Supports: .txt .pdf .docx .xlsx .csv .json")
        sys.exit(1)

    path = sys.argv[1]
    if not os.path.exists(path): print(f"File not found: {path}"); sys.exit(1)
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    print("="*50)
    print("  TEMP LABOR RFP GENERATOR")
    print("="*50)
    print(f"\n  Input: {path}\n")

    content = read_file(path)
    print(f"  Read: {len(content):,} chars")

    D = call_ai(content)

    ps = D.get('projectSetup',{})
    reqs = D.get('requirements',{}).get('items',[])
    qs = sum(len(s.get('questions',[])) for s in D.get('questionnaire',{}).get('sections',[]))
    roles = D.get('bidSheets',{}).get('roles',[])
    print(f"\n  Client: {ps.get('clientName','')}")
    print(f"  Requirements: {len(reqs)}")
    print(f"  Questions: {qs}")
    print(f"  SLAs: {len(D.get('slas',[]))}")
    print(f"  Roles: {len(roles)}")
    print(f"  Mark-up rows: {len(D.get('markup',{}).get('rows',[]))}")

    client_name = ps.get('clientName','Client').replace(' ','_').replace('/','_')
    ts = datetime.now().strftime('%Y%m%d_%H%M')
    out = os.path.join(OUTPUT_DIR, f"{client_name}_Temp_Labor_RFP_{ts}.xlsx")

    print(f"\n  Building Excel...")
    build_excel(D, out)

    print(f"\n  DONE: {out}")
    print(f"  Size: {os.path.getsize(out):,} bytes")
    print("="*50)

    try: os.startfile(out)
    except: pass

if __name__ == '__main__': main()
