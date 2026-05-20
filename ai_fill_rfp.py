"""
AI Auto-Fill RFP
================
Takes any client document (text, PDF, Word, Excel, email dump) and uses Claude
to generate a complete structured RFP dataset that auto-populates all dashboard fields.

Usage:
    python ai_fill_rfp.py <input_file> [output.json]

Input: Any file with client info (meeting notes, prior RFP, RFI response, email, SOW, etc.)
Output: Structured JSON that the dashboard imports directly + Excel RFP file.
"""

import json
import sys
import os
from datetime import datetime

import anthropic

OUTPUT_DIR = r"c:\Isha\Claude\Temp labor\output"

# ===== CONFIGURE YOUR MODEL HERE =====
# Change this to whatever model name your corporate API proxy supports.
# Run: python -c "import anthropic; print(anthropic.Anthropic().base_url)" to check your endpoint.
# Common options: "claude-sonnet-4-20250514", "claude-3-5-sonnet", "claude-3-sonnet", "anthropic.claude-v2"
MODEL = os.environ.get("CLAUDE_MODEL", "claude-sonnet-4-20250514")

SYSTEM_PROMPT = """You are a Senior Procurement RFP Specialist. Given ANY document about a client's temporary labor needs (meeting notes, prior RFP, RFI response, emails, SOW, contract, or even rough bullet points), you generate a COMPLETE structured RFP dataset.

You must output ONLY a valid JSON object (no markdown fences, no explanation) with this exact structure:

{
  "projectSetup": {
    "clientName": "extracted or inferred company name",
    "projectName": "e.g., 2026 Temp Labor Program",
    "rfpTitle": "Temporary Labor Global Agency RFP",
    "category": "Temporary Labor",
    "industry": "e.g., Pharmaceuticals",
    "description": "1-2 paragraph company description",
    "contactName": "extracted contact or leave blank",
    "contactEmail": "extracted or leave blank",
    "contactPhone": "",
    "procurementLead": "",
    "procurementLeadEmail": "",
    "timeline": {
      "issueDate": "YYYY-MM-DD or blank",
      "questionDeadline": "YYYY-MM-DD or blank",
      "proposalDue": "YYYY-MM-DD or blank",
      "awardDate": "",
      "startDate": ""
    },
    "duration": "e.g., 3 years",
    "currency": "USD",
    "countries": "comma-separated list of countries in scope",
    "states": "",
    "units": "business units/regions",
    "spend": "estimated annual spend if mentioned",
    "headcount": "estimated headcount if mentioned",
    "hours": "",
    "incumbentExists": true/false,
    "incumbents": "known incumbent suppliers, one per line",
    "paymentTerms": "Net 60",
    "notes": ""
  },
  "suppliers": [
    {"name": "supplier name", "type": "Incumbent|New|Preferred|Diverse", "contact": "", "email": "", "phone": "", "regions": "", "diversity": "", "notes": "", "invited": true}
  ],
  "requirements": {
    "overview": "2-3 paragraph program overview tailored to this client",
    "objectives": "bullet points of business objectives, one per line",
    "laborCategories": [
      {"title": "role title", "family": "job family", "skill": "Entry|Intermediate|Senior|Specialized", "work": "Onsite|Hybrid|Remote", "headcount": "number", "hours": "", "desc": "brief description"}
    ],
    "locations": [
      {"country": "country", "state": "", "city": "", "site": "site name if known", "onsite": true}
    ],
    "items": [
      {"category": "All|Operations|Sterile|Non-Sterile|Commercial|HR/Benefits|Professional Services", "subcategory": "e.g., Workforce Readiness & Onboarding", "text": "specific, measurable requirement text"}
    ]
  },
  "questionnaire": {
    "sections": [
      {
        "id": "unique_id",
        "name": "section name",
        "questions": [
          {"text": "question text", "type": "Text|Yes/No|Number", "weight": "1", "required": true}
        ]
      }
    ]
  },
  "slas": [
    {"name": "KPI name", "target": "e.g., ≥90%", "description": "what it measures", "metric": "Percentage|Time|Count", "frequency": "Monthly|Quarterly", "source": "measurement source", "penalty": "remedy if failed"}
  ],
  "bidSheets": {
    "model": "payrate_markup",
    "rateType": "hourly",
    "currency": "USD",
    "includeOT": true,
    "includeDT": false,
    "includeHoliday": false,
    "includeShift": false,
    "instructions": "clear instructions for suppliers on how to fill bid sheets",
    "roles": [
      {"id": "XX0001", "region": "APAC|EMEA|LATAM|NA", "family": "job family", "title": "job title", "country": "country", "site": "city/site", "demand": "number", "rate": "reference rate USD/hr", "desc": "role description"}
    ]
  },
  "markup": {
    "approach": "by_location_family",
    "defaultPct": "",
    "statutory": true,
    "fringe": true,
    "operating": true,
    "serviceFee": true,
    "instructions": "mark-up instructions for suppliers",
    "rows": [
      {"region": "APAC|EMEA|LATAM|North America", "location": "city, country", "country": "country", "family": "job family", "demand": "number"}
    ]
  },
  "discounts": {
    "convFeeNote": "$0 unless stated otherwise",
    "conversion": [],
    "rebate": [],
    "volume": [],
    "tenure": []
  },
  "conversion": {
    "allowed": true,
    "structure": "tiered",
    "guarantee": "",
    "buyout": "",
    "instructions": "",
    "feeRows": [
      {"period": "0-30 days", "type": "Percentage", "fee": "15%", "notes": ""}
    ],
    "fxRates": [
      {"location": "city, country", "country": "country", "currency": "Currency Name (CODE)", "rate": "rate to 1 USD"}
    ]
  }
}

RULES:
1. Extract ALL available info from the input. If something is mentioned, use it.
2. For anything NOT mentioned, generate industry-appropriate defaults:
   - 50-80 requirements covering workforce readiness, staffing controls, compliance, reporting, health/safety, commercial management, governance
   - 8-10 questionnaire sections with 4-6 questions each (account mgmt, compliance, fulfillment, technology, pricing, service delivery, diversity, references)
   - 15-25 SLAs (fill rate ≥90%, time-to-fill ≤14 days, no-show ≤3%, invoice accuracy ≥98%, etc.)
   - Roles based on mentioned positions/locations with market-rate pay
   - Mark-up rows for each unique region+location+family combo
   - Currency conversion rates for non-USD countries
3. Requirements must be SPECIFIC and MEASURABLE - never vague
4. Pay rates should be realistic market rates for the country/role
5. If the document mentions specific people, dates, locations, headcounts - USE THEM
6. Generate role IDs as region_prefix + C + 4-digit number (AC0001 for APAC, EC0001 for EMEA, etc.)
7. Include conversion fee tiers (0-30 days, 31-60 days, 61-90 days, 90+ days typical structure)
8. Always include fxRates for non-USD countries in scope

Output ONLY the JSON. No markdown, no explanation, no preamble."""


def read_file(filepath):
    """Read any file type and return text content."""
    ext = os.path.splitext(filepath)[1].lower()

    if ext == '.json':
        with open(filepath, 'r', encoding='utf-8') as f:
            return f.read()

    elif ext in ('.txt', '.csv', '.md'):
        with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
            return f.read()

    elif ext in ('.xlsx', '.xls'):
        from openpyxl import load_workbook
        wb = load_workbook(filepath, data_only=True)
        parts = []
        for name in wb.sheetnames:
            ws = wb[name]
            parts.append(f"\n=== Sheet: {name} ===")
            for row in ws.iter_rows(values_only=True):
                vals = [str(c).strip() for c in row if c]
                if vals:
                    parts.append(' | '.join(vals))
        return '\n'.join(parts)

    elif ext == '.docx':
        try:
            from docx import Document
        except ImportError:
            os.system("pip install python-docx --quiet")
            from docx import Document
        doc = Document(filepath)
        return '\n'.join([p.text for p in doc.paragraphs if p.text.strip()])

    elif ext == '.pdf':
        try:
            import PyPDF2
        except ImportError:
            os.system("pip install PyPDF2 --quiet")
            import PyPDF2
        with open(filepath, 'rb') as f:
            reader = PyPDF2.PdfReader(f)
            return '\n'.join([page.extract_text() or '' for page in reader.pages])

    else:
        with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
            return f.read()


def ai_generate(content):
    """Call Claude to generate the structured RFP data."""
    import httpx

    # Handle corporate SSL (self-signed certs)
    http_client = httpx.Client(verify=False)
    client = anthropic.Anthropic(http_client=http_client)

    # Truncate very long documents to stay within limits
    if len(content) > 80000:
        content = content[:80000] + "\n\n[Document truncated for processing]"

    print("  Calling Claude API...")
    response = client.messages.create(
        model=MODEL,
        max_tokens=16000,
        system=SYSTEM_PROMPT,
        messages=[{"role": "user", "content": f"Generate a complete Temp Labor RFP dataset from the following client document:\n\n{content}"}]
    )

    raw = response.content[0].text.strip()
    # Clean markdown fences if present
    if raw.startswith("```"):
        raw = raw.split("\n", 1)[1]
    if raw.endswith("```"):
        raw = raw.rsplit("```", 1)[0]
    if raw.startswith("json"):
        raw = raw[4:].strip()

    return json.loads(raw)


def generate_excel(data):
    """Generate the Excel RFP workbook directly."""
    try:
        import subprocess
        # Write temp JSON, call rfp_generator.py
        temp_json = os.path.join(OUTPUT_DIR, '_temp_ai_fill.json')
        with open(temp_json, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False)

        # Use the rfp_generator to produce formatted Excel
        result = subprocess.run(
            [sys.executable, os.path.join(os.path.dirname(__file__), 'rfp_generator.py'), temp_json],
            capture_output=True, text=True
        )
        if result.returncode == 0:
            print(result.stdout)
        else:
            print("  Excel generation via template skipped (using SheetJS in browser instead)")
            print(f"  {result.stderr[:200]}" if result.stderr else "")

        # Clean up temp
        try:
            os.remove(temp_json)
        except:
            pass
    except Exception as e:
        print(f"  Note: Template-based Excel generation skipped: {e}")


def main():
    if len(sys.argv) < 2:
        print("=" * 55)
        print("  AI AUTO-FILL RFP")
        print("=" * 55)
        print()
        print("  Usage: python ai_fill_rfp.py <client_file>")
        print()
        print("  Supported: .pdf .docx .xlsx .txt .csv .json .md")
        print()
        print("  Reads your client document, AI generates the full")
        print("  RFP dataset, outputs JSON for dashboard import +")
        print("  Excel RFP file.")
        sys.exit(1)

    filepath = sys.argv[1]
    if not os.path.exists(filepath):
        print(f"ERROR: File not found: {filepath}")
        sys.exit(1)

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    print("=" * 55)
    print("  AI AUTO-FILL RFP")
    print("=" * 55)
    print()
    print(f"  Input: {filepath}")
    print()

    # Step 1: Read file
    print("  Reading file...")
    content = read_file(filepath)
    print(f"  Content: {len(content):,} characters")
    print()

    # Step 2: AI generates structured data
    data = ai_generate(content)

    # Step 3: Count what was generated
    ps = data.get('projectSetup', {})
    reqs = data.get('requirements', {}).get('items', [])
    questions = sum(len(s.get('questions', [])) for s in data.get('questionnaire', {}).get('sections', []))
    slas = data.get('slas', [])
    roles = data.get('bidSheets', {}).get('roles', [])
    markups = data.get('markup', {}).get('rows', [])

    print()
    print(f"  Client: {ps.get('clientName', 'Unknown')}")
    print(f"  Industry: {ps.get('industry', 'N/A')}")
    print(f"  Countries: {ps.get('countries', 'N/A')}")
    print(f"  Requirements: {len(reqs)}")
    print(f"  Questions: {questions}")
    print(f"  SLAs: {len(slas)}")
    print(f"  Roles: {len(roles)}")
    print(f"  Mark-up rows: {len(markups)}")
    print()

    # Step 4: Save JSON for dashboard import
    client_name = ps.get('clientName', 'Client').replace(' ', '_').replace('/', '_')
    timestamp = datetime.now().strftime('%Y%m%d_%H%M')
    json_path = os.path.join(OUTPUT_DIR, f"{client_name}_rfp_autofill_{timestamp}.json")
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

    print(f"  JSON saved: {json_path}")
    print(f"  → Import this into the dashboard via 'Import File' button")
    print()

    # Step 5: Also generate Excel directly
    print("  Generating Excel...")
    generate_excel(data)

    print()
    print("=" * 55)
    print("  DONE!")
    print(f"  Import file: {json_path}")
    print("=" * 55)

    # Try to open the output folder
    try:
        os.startfile(OUTPUT_DIR)
    except:
        pass

    return json_path


if __name__ == '__main__':
    main()
