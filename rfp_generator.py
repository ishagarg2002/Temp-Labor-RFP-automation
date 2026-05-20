"""
Client-Specific Temp Labor RFP Generator
==========================================
Copies the reference template for formatting ONLY, clears all sample data,
and fills with client-specific data from the dashboard JSON export.

Usage:
    python rfp_generator.py <client_rfp_data.json>

Output: Excel file in ./output/ matching template formatting exactly.
"""

import json
import sys
import os
import shutil
from copy import copy
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

TEMPLATE_PATH = r"C:\Users\igarg015\Downloads\Shared_Viatris_ Temporary Labor_Agency_RFX (1).xlsx"
OUTPUT_DIR = r"c:\Isha\Claude\Temp labor\output"


def safe_write(ws, row, col, value):
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        for mr in list(ws.merged_cells.ranges):
            if cell.coordinate in mr:
                ws.unmerge_cells(str(mr))
                break
        cell = ws.cell(row=row, column=col)
    cell.value = value
    return cell


def safe_clear(ws, row, col):
    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        cell.value = None


def get_row_styles(ws, row, max_col):
    styles = {}
    for col in range(1, max_col):
        cell = ws.cell(row=row, column=col)
        if not isinstance(cell, MergedCell):
            styles[col] = {'font': copy(cell.font), 'fill': copy(cell.fill), 'border': copy(cell.border), 'alignment': copy(cell.alignment)}
    return styles


def apply_styles(ws, row, styles, max_col):
    for col in range(1, max_col):
        cell = ws.cell(row=row, column=col)
        if isinstance(cell, MergedCell):
            continue
        if col in styles:
            cell.font = copy(styles[col]['font'])
            cell.fill = copy(styles[col]['fill'])
            cell.border = copy(styles[col]['border'])
            cell.alignment = copy(styles[col]['alignment'])


def find_header_row(ws, text, col=2, max_row=30):
    for r in range(1, max_row):
        v = ws.cell(row=r, column=col).value
        if v and text in str(v):
            return r
    return None


def clear_data_rows(ws, start_row, max_col=14):
    for r in range(start_row, ws.max_row + 1):
        for c in range(1, max_col):
            safe_clear(ws, r, c)


def detect_model_version(data):
    """Detect if this is the new v4 model (has projectSetup) or old flat model."""
    return 'projectSetup' in data


def generate_rfp(data):
    """Generate client-specific RFP Excel from dashboard export data."""

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # Detect model version
    is_v4 = detect_model_version(data)

    if is_v4:
        ps = data['projectSetup']
        client = ps.get('clientName', 'Client')
    else:
        ps = data.get('setup', {})
        client = ps.get('company', 'Client')

    timestamp = datetime.now().strftime('%Y%m%d_%H%M')
    safe_client = client.replace(' ', '_').replace('/', '_')
    output_path = os.path.join(OUTPUT_DIR, f"{safe_client}_Temp_Labor_RFP_{timestamp}.xlsx")

    # Copy template (formatting source only)
    shutil.copy2(TEMPLATE_PATH, output_path)
    wb = load_workbook(output_path)

    print(f"Generating RFP for: {client}")
    print(f"Template: {TEMPLATE_PATH}")
    print()

    # ===== TAB 1: COVER LETTER =====
    ws = wb['1. Cover Letter']
    # Clear all content cells (keep structure)
    for r in range(1, ws.max_row + 1):
        for c in range(2, 4):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell, MergedCell):
                continue
            if cell.value and isinstance(cell.value, str):
                # Replace Viatris references
                if 'Viatris' in cell.value:
                    cell.value = cell.value.replace('Viatris', client).replace('viatris', client.lower())

    # Set title
    ws.cell(row=2, column=2).value = f"{client} {ps.get('rfpTitle', 'Temporary Labor Global Agency RFP')}" if is_v4 else f"{client} Temporary Labor RFP"

    # Set contact
    for r in range(1, ws.max_row + 1):
        cell = ws.cell(row=r, column=2)
        if cell.value and isinstance(cell.value, str):
            if 'Point of Contact' in cell.value:
                ws.cell(row=r, column=2).value = f"{client} Point of Contact"
            if 'Name:' in str(cell.value):
                cell.value = f"Name: {ps.get('contactName', '') if is_v4 else ps.get('contact_name', '')}"
            if 'Email:' in str(cell.value):
                cell.value = f"Email: {ps.get('contactEmail', '') if is_v4 else ps.get('contact_email', '')}"

    print("  [1] Cover Letter - updated")

    # ===== TAB 3: REQUIREMENTS =====
    ws = wb['3. Requirements']
    items = data.get('requirements', {}).get('items', []) if is_v4 else data.get('requirements', [])

    # Clear all data rows
    header_row = find_header_row(ws, '#', col=2)
    if header_row:
        start = header_row + 1
        styles = get_row_styles(ws, start, 9)
        clear_data_rows(ws, start, 9)

        for i, req in enumerate(items):
            r = start + i
            safe_write(ws, r, 2, i + 1)
            safe_write(ws, r, 3, req.get('category', 'All'))
            safe_write(ws, r, 4, req.get('subcategory', ''))
            safe_write(ws, r, 5, req.get('text', ''))
            safe_write(ws, r, 6, '')  # Response (supplier fills)
            safe_write(ws, r, 7, '')  # Comment (supplier fills)
            apply_styles(ws, r, styles, 9)

    print(f"  [3] Requirements - {len(items)} items")

    # ===== TAB 4: QUESTIONNAIRE =====
    ws = wb['4. Questionnaire']
    if is_v4:
        questions = []
        for sec in data.get('questionnaire', {}).get('sections', []):
            for q in sec.get('questions', []):
                questions.append({'category': sec['name'], 'text': q['text']})
    else:
        questions = data.get('questions', [])

    header_row = find_header_row(ws, '#', col=2)
    if header_row:
        start = header_row + 1
        styles = get_row_styles(ws, start, 7)
        clear_data_rows(ws, start, 7)

        for i, q in enumerate(questions):
            r = start + i
            safe_write(ws, r, 2, i + 1)
            safe_write(ws, r, 3, q.get('category', 'General'))
            safe_write(ws, r, 4, q.get('text', ''))
            safe_write(ws, r, 5, '')  # Response (supplier fills)
            apply_styles(ws, r, styles, 7)

    print(f"  [4] Questionnaire - {len(questions)} questions")

    # ===== TAB 5: SLAs =====
    ws = wb['5. Service Level Agreements']
    slas = data.get('slas', [])

    # Unmerge all to avoid write errors
    for mr in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(mr))

    header_row = find_header_row(ws, 'KPI', col=2)
    if header_row:
        start = header_row + 1
        clear_data_rows(ws, start, 10)

        for i, sla in enumerate(slas):
            r = start + i
            ws.cell(row=r, column=2).value = sla.get('name', sla.get('kpi', ''))
            ws.cell(row=r, column=3).value = sla.get('description', sla.get('definition', ''))
            ws.cell(row=r, column=4).value = sla.get('calculation', sla.get('calc', ''))
            ws.cell(row=r, column=5).value = sla.get('target', '')
            ws.cell(row=r, column=6).value = sla.get('frequency', sla.get('period', 'Monthly'))
            ws.cell(row=r, column=7).value = sla.get('source', '')
            ws.cell(row=r, column=8).value = ''  # Comply (supplier)
            ws.cell(row=r, column=9).value = ''  # Comments (supplier)

    print(f"  [5] SLAs - {len(slas)} KPIs")

    # ===== TABS 6a-6d: BID SHEETS =====
    region_map = {'APAC': '6.a Bid Sheet_APAC', 'EMEA': '6.b Bid Sheet_EMEA', 'LATAM': '6.c Bid Sheet_LATAM', 'NA': '6.d Bid Sheet_NA'}
    roles = data.get('bidSheets', {}).get('roles', []) if is_v4 else data.get('roles', [])

    for region, sheet_name in region_map.items():
        ws = wb[sheet_name]
        region_roles = [r for r in roles if r.get('region') == region]

        # Replace Viatris in title
        for r in range(1, 5):
            for c in range(2, 14):
                cell = ws.cell(row=r, column=c)
                if isinstance(cell, MergedCell):
                    continue
                if cell.value and isinstance(cell.value, str) and 'Viatris' in cell.value:
                    cell.value = cell.value.replace('Viatris', client)

        header_row = find_header_row(ws, 'RFP ID', col=2)
        if not header_row:
            continue
        start = header_row + 2  # skip header + example row
        styles = get_row_styles(ws, start - 1, 14)
        clear_data_rows(ws, start, 14)

        for i, role in enumerate(region_roles):
            r = start + i
            safe_write(ws, r, 2, role.get('id', role.get('rfp_id', f'{region[0]}C{str(i+1).zfill(4)}')))
            safe_write(ws, r, 3, role.get('family', role.get('job_family', '')))
            safe_write(ws, r, 4, role.get('title', role.get('job_title', '')))
            safe_write(ws, r, 5, role.get('desc', role.get('description', '')))
            safe_write(ws, r, 6, role.get('country', ''))
            safe_write(ws, r, 7, role.get('site', ''))
            safe_write(ws, r, 8, role.get('demand', ''))
            safe_write(ws, r, 9, role.get('rate', role.get('pay_rate', '')))
            safe_write(ws, r, 10, '')  # Proposed (supplier)
            safe_write(ws, r, 11, '')  # OT (supplier)
            safe_write(ws, r, 12, '')  # Bidding (supplier)
            safe_write(ws, r, 13, '')  # Comments (supplier)
            apply_styles(ws, r, styles, 14)

        print(f"  [6] {sheet_name} - {len(region_roles)} roles")

    # ===== TAB 7: MARK-UP =====
    ws = wb['7. Bid Sheet_Mark-Up']
    markups = data.get('markup', {}).get('rows', []) if is_v4 else data.get('markups', [])

    # Replace Viatris
    for r in range(1, 5):
        for c in range(2, 10):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell, MergedCell):
                continue
            if cell.value and isinstance(cell.value, str) and 'Viatris' in cell.value:
                cell.value = cell.value.replace('Viatris', client)

    header_row = find_header_row(ws, 'Region', col=2)
    if header_row:
        start = header_row + 2  # skip header + example
        styles = get_row_styles(ws, start - 1, 10)
        clear_data_rows(ws, start, 10)

        for i, mu in enumerate(markups):
            r = start + i
            safe_write(ws, r, 2, mu.get('region', ''))
            safe_write(ws, r, 3, mu.get('location', mu.get('loc', '')))
            safe_write(ws, r, 4, mu.get('country', ''))
            safe_write(ws, r, 5, mu.get('family', mu.get('job_family', '')))
            safe_write(ws, r, 6, mu.get('demand', ''))
            safe_write(ws, r, 7, '')  # Standard % (supplier)
            safe_write(ws, r, 8, '')  # OT % (supplier)
            safe_write(ws, r, 9, '')  # Comments (supplier)
            apply_styles(ws, r, styles, 10)

    print(f"  [7] Mark-Up - {len(markups)} rows")

    # ===== TAB 8: DISCOUNTS =====
    ws = wb['8. Discounts']

    # Replace Viatris in instructions
    for r in range(1, ws.max_row + 1):
        for c in range(1, 10):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell, MergedCell):
                continue
            if cell.value and isinstance(cell.value, str) and 'Viatris' in cell.value:
                cell.value = cell.value.replace('Viatris', client)

    print(f"  [8] Discounts - client name updated")

    # ===== TAB 9: CONVERSION RATES =====
    ws = wb['9. Conversion Rates']
    fx_rates = data.get('conversion', {}).get('fxRates', []) if is_v4 else data.get('conversions', [])

    # Replace Viatris
    for r in range(1, 5):
        for c in range(2, 10):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell, MergedCell):
                continue
            if cell.value and isinstance(cell.value, str) and 'Viatris' in cell.value:
                cell.value = cell.value.replace('Viatris', client)

    if fx_rates:
        start = 6  # data starts at row 6 in template
        clear_data_rows(ws, start, 7)

        for i, fx in enumerate(fx_rates):
            r = start + i
            safe_write(ws, r, 3, fx.get('location', fx.get('loc', '')))
            safe_write(ws, r, 4, fx.get('country', ''))
            safe_write(ws, r, 5, fx.get('currency', ''))
            try:
                safe_write(ws, r, 6, float(fx.get('rate', 0)))
            except (ValueError, TypeError):
                safe_write(ws, r, 6, fx.get('rate', ''))

    print(f"  [9] Conversion - {len(fx_rates)} rates")

    # ===== TAB 2: CONTENTS (update counts) =====
    ws = wb['2. Contents & Response Summary']
    # Replace Viatris
    for r in range(1, ws.max_row + 1):
        for c in range(1, 22):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell, MergedCell):
                continue
            if cell.value and isinstance(cell.value, str) and 'Viatris' in cell.value:
                cell.value = cell.value.replace('Viatris', client)

    print(f"  [2] Contents - client name updated")

    # Save
    wb.save(output_path)
    print()
    print("=" * 50)
    print(f"DONE: {output_path}")
    print(f"Size: {os.path.getsize(output_path):,} bytes")
    print("=" * 50)

    # Open on Windows
    try:
        os.startfile(output_path)
    except:
        pass

    return output_path


def main():
    if len(sys.argv) < 2:
        print("Usage: python rfp_generator.py <client_rfp_data.json>")
        print()
        print("Export JSON from the RFP Builder dashboard, then run this script.")
        print("Output Excel will match the template formatting with client-specific data.")
        sys.exit(1)

    filepath = sys.argv[1]
    if not os.path.exists(filepath):
        print(f"ERROR: File not found: {filepath}")
        sys.exit(1)

    with open(filepath, 'r', encoding='utf-8') as f:
        data = json.load(f)

    generate_rfp(data)


if __name__ == '__main__':
    main()
